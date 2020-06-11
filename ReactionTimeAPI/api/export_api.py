from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_404_NOT_FOUND
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from openpyxl import Workbook

from ..models import Experiment, ExperimentResult, Score

header_font = Font(bold=True)

centered_alignment = Alignment(horizontal='center')

border_bottom = Border(
    bottom=Side(border_style='medium', color='FF000000'),
)

border_bottom_thin = Border(
    bottom=Side(border_style='thin', color='FF000000'),
)

wrapped_alignment = Alignment(
    vertical='top',
    wrap_text=True
)

fill = PatternFill(
    start_color='FFD3D3D3',
    end_color='FFD3D3D3',
    fill_type='solid',
)

fill_light = PatternFill(
    start_color='FFF5F5F5',
    end_color='FFF5F5F5',
    fill_type='solid',
)


@api_view(['POST'])
@authentication_classes((TokenAuthentication,))
@permission_classes((IsAuthenticated, IsAdminUser,))
def export_single_experiment_results(request, pk):
    """
    API endpoint for exporting experiment results in XLSX with provided parameters.

    Returns:

    """

    date_from = request.data.get("dateFrom")
    date_to = request.data.get("dateTo")

    try:
        experiment = Experiment.objects.get(pk=pk)
    except Experiment.DoesNotExist:
        return Response('Experiment with given ID not found', status=HTTP_404_NOT_FOUND)

    if date_from and date_to:
        results = ExperimentResult.objects.filter(experiment=experiment,
                                                  date__gte=date_from,
                                                  date__lte=date_to
                                                  )
    elif date_from:
        results = ExperimentResult.objects.filter(experiment=experiment, date__gte=date_from)
    elif date_to:
        results = ExperimentResult.objects.filter(experiment=experiment, date__lte=date_to)
    else:
        results = ExperimentResult.objects.filter(experiment=experiment)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Results'

    # Define the titles for columns
    columns = [
        ('ID', 5),
        ('First Name', 15),
        ('Last Name', 15),
        ('Email', 25),
        ('Date', 20),
        ('Score Type', 15),
        ('Average', 10),
        ('Best', 10),
        ('Success', 10),
        ('Required data section →', 25)
    ]
    req_data_separator_column = len(columns)

    required_data_header = []
    for res in results:
        keys = res.requiredData.keys()
        for key in keys:
            if not any(elem[0] == key for elem in required_data_header):
                required_data_header.append((key, 20))

    columns.extend(required_data_header)

    # Assign the titles for each cell of the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        if col_num == req_data_separator_column:
            cell.fill = fill

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    row_num = 1
    for res in results:
        scores = Score.objects.filter(experimentResult=res)

        for score in scores:
            row_num += 1

            # Define the data for each cell in the row
            row = [None] * len(columns)

            row[0] = res.id
            row[1] = res.first_name
            row[2] = res.last_name
            row[3] = res.email
            row[4] = res.date
            row[5] = score.type
            row[6] = score.average
            row[7] = score.best
            row[8] = score.success

            columns_titles = list(col[0] for col in columns[(req_data_separator_column-1):])
            for key, value in res.requiredData.items():
                if key in columns_titles:
                    idx = columns_titles.index(key)
                    row[req_data_separator_column-1 + idx] = value

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if col_num == req_data_separator_column:
                    cell.fill = fill

    worksheet.freeze_panes = worksheet['A2']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        status=HTTP_200_OK,
    )
    response['Content-Disposition'] = 'attachment; filename={date}-export-{id}.xlsx'.format(
        date=timezone.now().strftime('%Y-%m-%d'), id=experiment.pk
    )
    workbook.save(response)

    return response


@api_view(['POST'])
@authentication_classes((TokenAuthentication,))
@permission_classes((IsAuthenticated, IsAdminUser,))
def export_all_results(request):
    """
    API endpoint for exporting experiment results in XLSX with provided parameters.

    Returns:

    """

    expiration_from = request.data.get("expirationFrom")
    expiration_to = request.data.get("expirationTo")
    created_from = request.data.get("createdFrom")
    created_to = request.data.get("createdTo")
    allow_multiple_answers = request.data.get("allowMultipleAnswers")

    experiments = Experiment.objects.all()
    if expiration_from:
        experiments = experiments.filter(expiration__gte=expiration_from)
    if expiration_to:
        experiments = experiments.filter(expiration__lte=expiration_to)
    if created_from:
        experiments = experiments.filter(created__gte=created_from)
    if created_to:
        experiments = experiments.filter(created__lte=created_to)
    if allow_multiple_answers:
        experiments = experiments.filter(allowMultipleAnswers=allow_multiple_answers)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Results'

    # Define the titles for columns
    columns = [
        ('Exp ID', 8),
        ('Exp Created', 20),
        ('Exp Expiration', 20),
        ('ID', 5),
        ('First Name', 15),
        ('Last Name', 15),
        ('Email', 25),
        ('Date', 20),
        ('Score Type', 15),
        ('Average', 10),
        ('Best', 10),
        ('Success', 10),
        ('Required data section →', 25)
    ]
    req_data_separator_column = len(columns)
    exp_data_last_column = 3
    required_data_header = []

    for exp in experiments:
        results = ExperimentResult.objects.filter(experiment=exp)
        for res in results:
            keys = res.requiredData.keys()
            for key in keys:
                if not any(elem[0] == key for elem in required_data_header):
                    required_data_header.append((key, 20))

    columns.extend(required_data_header)

    # Assign the titles for each cell of the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment

        if col_num <= exp_data_last_column:
            cell.fill = fill_light

        if col_num == req_data_separator_column:
            cell.fill = fill

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    row_num = 1
    for exp in experiments:
        results = ExperimentResult.objects.filter(experiment=exp)
        for res in results:
            scores = Score.objects.filter(experimentResult=res)

            for score in scores:
                row_num += 1

                # Define the data for each cell in the row
                row = [None] * len(columns)

                row[0] = exp.id
                row[1] = exp.created
                row[2] = exp.expiration
                row[3] = res.id
                row[4] = res.first_name
                row[5] = res.last_name
                row[6] = res.email
                row[7] = res.date
                row[8] = score.type
                row[9] = score.average
                row[10] = score.best
                row[11] = score.success

                columns_titles = list(col[0] for col in columns[(req_data_separator_column-1):])
                for key, value in res.requiredData.items():
                    if key in columns_titles:
                        idx = columns_titles.index(key)
                        row[req_data_separator_column-1 + idx] = value

                # Assign the data for each cell of the row
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

                    if col_num == 1:
                        cell.font = header_font

                    if col_num <= exp_data_last_column:
                        cell.fill = fill_light

                    if col_num == req_data_separator_column:
                        cell.fill = fill

                    if list(results).index(res) == (len(results) - 1) \
                            and list(scores).index(score) == (len(scores) - 1):
                        cell.border = border_bottom_thin

    worksheet.freeze_panes = worksheet['A2']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        status=HTTP_200_OK
    )
    response['Content-Disposition'] = 'attachment; filename={date}-export.xlsx'.format(
        date=timezone.now().strftime('%Y-%m-%d'),
    )
    workbook.save(response)

    return response
